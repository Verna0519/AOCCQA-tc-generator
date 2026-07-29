#!/usr/bin/env python3
"""
AOCCQA-case-exporter  (v1.2.0 — keep in sync with SKILL.md metadata.version)

Fill the AOCC QA xlsx template from:
  (1) a Jira ticket's fields  -> Report sheet + filename
  (2) the previous agent's test cases (7-col standard) -> Test case sheet

Deterministic formatting only. No content judgement, filtering, or rewriting.
Bug list / Screenshot sheets and all formulas are left untouched.
"""

import argparse
import copy
import json
import os
import re
import sys
from datetime import datetime

import openpyxl

MAX_CASES = 200          # template rows 2..201
FIRST_DATA_ROW = 2

# Report sheet: label cell (col A) -> input cell (col C)
REPORT_CELLS = {
    "summary": "C2",         # Project cell <- full Jira Summary (tags kept)
    "test_date": "C3",
    "test_version": "C4",
    "tester": "C5",
    "link": "C6",            # New feature & Release Note
    "mcc": "C13",            # Test Country
    "test_environment": "C14",
}

# Test case sheet: previous-agent field -> template column letter
CASE_COLS = {
    "id": "A",
    "category": "E",
    "pre_condition": "F",
    "test_case": "G",
    "steps": "H",
    "expected_result": "I",
    "test_data": "L",
}

ILLEGAL_FILENAME = re.compile(r'[\\/:*?"<>|]')
UAT_QA_TAG = re.compile(r'\[UAT-QA\]', flags=re.IGNORECASE)
LEADING_TAG = re.compile(r'^\s*\[([^\]]+)\]\s*')   # e.g. "[EU] " -> "EU_"
DATE_TOKEN = re.compile(r'(\d{4})[/\-.](\d{1,2})[/\-.](\d{1,2})')  # YYYY/MM/DD, YYYY-MM-DD, YYYY.MM.DD


def normalize_test_date(raw):
    """
    Normalize a free-form schedule string into the Report's Test date format:
    ``YYYY/MM/DD-YYYY/MM/DD`` (earliest start ~ latest end), dropping wording
    such as "Internal Testing" / "UAT". Deterministic: parses whatever dates
    are literally present, never invents one.

    Returns (value, note):
      - value: normalized range, a single ``YYYY/MM/DD`` when only one distinct
               date is present, the stripped raw text when no date parses
               (so information is not silently lost), or "" for empty input.
      - note:  None on a clean normalization; otherwise a short reason string
               so the run report can flag it for a human to correct.
    """
    s = (raw or "").strip()
    if not s:
        return "", None
    found = []
    for y, m, d in DATE_TOKEN.findall(s):
        try:
            found.append(datetime(int(y), int(m), int(d)))
        except ValueError:
            pass  # e.g. 2026/13/40 — ignore impossible dates, never guess
    if not found:
        return s, "no parseable YYYY/MM/DD date; wrote raw text verbatim"
    lo, hi = min(found), max(found)
    if lo == hi:
        return lo.strftime("%Y/%m/%d"), None
    return f"{lo.strftime('%Y/%m/%d')}-{hi.strftime('%Y/%m/%d')}", None


def clean_summary_to_filename(summary: str, today: str, module: str = "") -> str:
    """
    Turn a Jira Summary into the export filename.

    Prefix segments are the leading bracket tags, in order, joined by "_".
    House style is a two-segment [market][module] prefix, but one segment
    (or three) still works — whatever the Summary carries is what is used.

      "[UAT-QA][TW][CRM] ASUS Membership ... threshold"
        -> remove [UAT-QA]            -> "[TW][CRM] ASUS Membership ... threshold"
        -> tags become prefix         -> "TW_CRM_ASUS Membership ... threshold"
        -> append suffix              -> "TW_CRM_ASUS Membership ... threshold_TestCase_YYYYMMDD.xlsx"

    When the Summary carries only the market tag, an explicit jira.module
    value is inserted as the second segment:

      summary "[UAT-QA][TW] ASUS Membership ..." + module "CRM"
        -> "TW_CRM_ASUS Membership ..._TestCase_YYYYMMDD.xlsx"

    A module already present in the Summary tags is never duplicated.
    """
    base = UAT_QA_TAG.sub("", summary or "").strip()

    # Consume every leading [tag] into an ordered prefix list.
    segments = []
    while True:
        m = LEADING_TAG.match(base)
        if not m:
            break
        seg = m.group(1).strip()
        if seg:
            segments.append(seg)
        base = base[m.end():].lstrip()

    module = (module or "").strip()
    if module and module.casefold() not in [s.casefold() for s in segments]:
        # market stays first; module becomes the second segment
        segments.insert(1, module) if segments else segments.append(module)

    title = base.strip()
    if not segments and not title:
        title = "Untitled"

    stem = "_".join(segments + ([title] if title else []))
    name = f"{stem}_TestCase_{today}.xlsx"
    name = ILLEGAL_FILENAME.sub("_", name)   # keep spaces/underscores as-is
    return name


def write_report(ws, jira: dict):
    """
    Fill Report input cells from Jira. Dynamic fields are filled only when a
    non-empty value is supplied; missing ones are left blank (never guessed).
    Returns a capture report listing captured vs blank for each dynamic field.
    """
    # dynamic Jira-sourced fields written verbatim: key -> (cell, human label)
    dynamic = {
        "summary":          (REPORT_CELLS["summary"],          "Project (Summary)"),
        "link":             (REPORT_CELLS["link"],             "New feature & Release Note (link)"),
        "mcc":              (REPORT_CELLS["mcc"],              "Test Country (MCC#)"),
        "test_environment": (REPORT_CELLS["test_environment"], "Test Environment"),
        "test_version":     (REPORT_CELLS["test_version"],     "Test Version"),
    }

    captured, blank = [], []
    for key, (cell, label) in dynamic.items():
        raw = jira.get(key)
        val = raw.strip() if isinstance(raw, str) else raw
        if val in (None, ""):
            blank.append({"field": label, "cell": cell})
        else:
            ws[cell] = val
            captured.append({"field": label, "cell": cell, "value": val})

    # Test date is normalized to YYYY/MM/DD-YYYY/MM/DD before writing (not verbatim).
    td_cell = REPORT_CELLS["test_date"]
    td_val, td_note = normalize_test_date(jira.get("test_date"))
    if td_val:
        entry = {"field": "Test date", "cell": td_cell, "value": td_val}
        if td_note:
            entry["note"] = td_note
        ws[td_cell] = td_val
        captured.append(entry)
    else:
        blank.append({"field": "Test date", "cell": td_cell})

    # Tester is derived, not raw: AOCC_<Assignee>. No assignee -> clear the cell
    # and report blank; never leave the template's pre-filled name on someone
    # else's deliverable (keeping a stray name is itself a wrong guess).
    assignee = (jira.get("assignee") or "").strip()
    tester_cell = REPORT_CELLS["tester"]
    if assignee:
        ws[tester_cell] = f"AOCC_{assignee}"
        captured.append({"field": "Tester", "cell": tester_cell, "value": f"AOCC_{assignee}"})
    else:
        ws[tester_cell] = None
        blank.append({"field": "Tester (no Assignee on ticket)", "cell": tester_cell})

    return captured, blank


def write_cases(ws, cases: list):
    """
    Write test cases into rows 2..201. Feature is intentionally dropped.

    Test Data -> column L is written only when the case supplies a non-empty
    value; otherwise L is left blank (never invented). Returns how many rows
    got Test Data, for the run report.
    """
    td_written = 0
    for i, case in enumerate(cases):
        row = FIRST_DATA_ROW + i
        cid = str(case.get("id") or (i + 1)).strip()
        ws[f"{CASE_COLS['id']}{row}"] = cid
        ws[f"{CASE_COLS['category']}{row}"] = case.get("category", "")
        ws[f"{CASE_COLS['pre_condition']}{row}"] = case.get("pre_condition", "")
        ws[f"{CASE_COLS['test_case']}{row}"] = case.get("test_case", "")
        ws[f"{CASE_COLS['steps']}{row}"] = case.get("steps", "")
        ws[f"{CASE_COLS['expected_result']}{row}"] = case.get("expected_result", "")

        raw_td = case.get("test_data")
        td = raw_td.strip() if isinstance(raw_td, str) else raw_td
        if td:
            ws[f"{CASE_COLS['test_data']}{row}"] = td
            td_written += 1
        # B/C/D (platform) and J/K (execution) left blank on purpose.
    return td_written


def main():
    ap = argparse.ArgumentParser(description="AOCCQA case exporter")
    ap.add_argument("--template", required=True, help="path to Test_Case_Template_Claude.xlsx")
    ap.add_argument("--input", required=True, help="path to input.json (jira + test_cases)")
    ap.add_argument("--outdir", default="/mnt/user-data/outputs", help="output directory")
    ap.add_argument("--date", default=None, help="override date as YYYYMMDD (default: today)")
    args = ap.parse_args()

    if not os.path.exists(args.template):
        sys.exit(f"ERROR: template not found: {args.template}")
    if not os.path.exists(args.input):
        sys.exit(f"ERROR: input.json not found: {args.input}")

    with open(args.input, encoding="utf-8") as f:
        data = json.load(f)

    jira = data.get("jira", {})
    cases = data.get("test_cases", [])

    if not cases:
        sys.exit("ERROR: 0 test cases supplied — nothing to export.")
    if len(cases) > MAX_CASES:
        sys.exit(f"ERROR: {len(cases)} cases exceed template limit of {MAX_CASES}.")

    today = args.date or datetime.now().strftime("%Y%m%d")

    wb = openpyxl.load_workbook(args.template)
    for required in ("Report", "Test case"):
        if required not in wb.sheetnames:
            sys.exit(
                f"ERROR: template is missing the required '{required}' sheet. "
                f"Found sheets: {wb.sheetnames}. Not building a substitute template."
            )
    report_ws = wb["Report"]
    case_ws = wb["Test case"]

    captured, blanks = write_report(report_ws, jira)
    td_written = write_cases(case_ws, cases)
    # Bug list / Screenshot untouched by design.

    fname = clean_summary_to_filename(jira.get("summary", ""), today, jira.get("module", ""))
    os.makedirs(args.outdir, exist_ok=True)
    out_path = os.path.join(args.outdir, fname)
    wb.save(out_path)

    result = {
        "output_path": out_path,
        "filename": fname,
        "case_count": len(cases),
        "test_data_filled": td_written,
        "test_data_blank": len(cases) - td_written,
        "report_captured": captured,
        "report_blank": blanks,
        "sheets_preserved": ["Bug list", "Screenshot"],
        "formulas_touched": False,
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
